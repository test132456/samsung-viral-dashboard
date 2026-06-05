"""'바이럴 일정.xlsx' (주간 달력 + 배포일정 레이아웃) → calendar_events / distribution 시트 적재.

사용:
   python scripts/import_viral_schedule.py "바이럴 일정.xlsx" --dry-run   # 파싱 결과만 출력
   python scripts/import_viral_schedule.py "바이럴 일정.xlsx"             # 구글시트에 적재
달력 영역은 cols 4~10(일~토), 배포일정 표는 cols 14~19로 가정한다.
"""
import sys, os, argparse, datetime
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from core import schema


def _is_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return True
    return isinstance(v, str) and len(v) >= 8 and v[:4].isdigit() and "-" in v[:10]


def _dstr(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()[:10]


def _track(task: str) -> str:
    return "공식" if "공식" in task else ("배포형" if "배포형" in task else "")


def parse(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    byrow = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                byrow.setdefault(r, {})[c] = v

    dow = {"일", "월", "화", "수", "목", "금", "토"}
    # 달력: cols 4~10에 날짜가 있는 행 = 주(週) 헤더
    date_rows = [r for r in sorted(byrow) if any(_is_date(byrow[r].get(c)) for c in range(4, 11))]
    events = []
    for i, r in enumerate(date_rows):
        nxt = date_rows[i + 1] if i + 1 < len(date_rows) else r + 6
        coldate = {c: _dstr(byrow[r][c]) for c in range(4, 11)
                   if c in byrow[r] and _is_date(byrow[r][c])}
        for tr in range(r + 2, nxt):  # r+1 = 요일행 skip
            for c, d in coldate.items():
                t = byrow.get(tr, {}).get(c)
                if t is None or _is_date(t):
                    continue
                t = str(t).replace("\n", " ").strip()
                if not t or t in dow:
                    continue
                events.append({"date": d, "task": t, "track": _track(t)})

    # 배포 일정: cols 14~19, '[...]' 또는 '공식블로그' 를 그룹 구분자로
    dist = []
    group = ""
    for r in sorted(byrow):
        row = byrow[r]
        head = str(row.get(14, "")).strip()
        if not head:
            continue
        if head.startswith("["):
            group = head.strip("[]").split("]")[0]
            continue
        if head in ("블로거",):
            continue
        if head == "공식블로그":
            group = "공식블로그"
            continue
        pub = byrow[r].get(15)
        pub_s = _dstr(pub) if _is_date(pub) else ""
        note = "" if _is_date(pub) or pub is None else str(pub).strip()
        dist.append({
            "group": group, "blogger": head, "publish_date": pub_s,
            "approval_no": str(row.get(16, "")).strip(),
            "landing_url": str(row.get(17, "")).strip(),
            "note": (note or str(row.get(18, "")).strip()),
            "publish_url": str(row.get(19, "")).strip(),
        })
    return events, dist


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    events, dist = parse(args.xlsx)
    print(f"달력 이벤트 {len(events)}건 / 배포 {len(dist)}건 파싱됨")
    if args.dry_run:
        for e in events:
            print("  [cal]", e["date"], e["track"], e["task"])
        for d in dist:
            print("  [dist]", d["group"], d["blogger"], d["publish_date"], d["approval_no"][:30])
        return

    import streamlit as st
    from core.sheets import Sheets
    info = dict(st.secrets["gcp_service_account"])
    sheets = Sheets(info, st.secrets["SPREADSHEET_ID"])
    sheets.ensure_tabs()
    for e in events:
        sheets.append(schema.SHEET_CALENDAR, e)
    for d in dist:
        sheets.append(schema.SHEET_DISTRIBUTION, d)
    print("적재 완료")


if __name__ == "__main__":
    main()
