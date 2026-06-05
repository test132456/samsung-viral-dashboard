"""'바이럴 일정.xlsx' (주간 달력 + 배포일정 레이아웃) 파서.

달력 영역 = cols 4~10 (일~토), 배포일정 표 = cols 14~19 로 가정.
parse(path) -> (events, distribution)
  events: [{date:'YYYY-MM-DD', task, track}]
  distribution: [{group, blogger, publish_date, approval_no, landing_url, note, publish_url}]
"""
from __future__ import annotations
import datetime
import openpyxl

_DOW = {"일", "월", "화", "수", "목", "금", "토"}


def _is_date(v) -> bool:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return True
    return isinstance(v, str) and len(v) >= 8 and v[:4].isdigit() and "-" in v[:10]


def _dstr(v) -> str:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()[:10]


def _track(task: str) -> str:
    return "공식" if "공식" in task else ("배포형" if "배포형" in task else "")


def parse(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    byrow: dict[int, dict] = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                byrow.setdefault(r, {})[c] = v

    # 달력: cols 4~10 에 날짜가 있는 행 = 주(週) 헤더
    date_rows = [r for r in sorted(byrow) if any(_is_date(byrow[r].get(c)) for c in range(4, 11))]
    events = []
    for i, r in enumerate(date_rows):
        nxt = date_rows[i + 1] if i + 1 < len(date_rows) else r + 6
        coldate = {c: _dstr(byrow[r][c]) for c in range(4, 11)
                   if c in byrow[r] and _is_date(byrow[r][c])}
        for tr in range(r + 2, nxt):  # r+1 = 요일행 skip
            for c, d in coldate.items():
                t = byrow.get(tr, {}).get(c)
                if t is None or _is_date(t):
                    continue
                t = str(t).replace("\n", " ").strip()
                if not t or t in _DOW:
                    continue
                events.append({"date": d, "task": t, "track": _track(t)})

    # 배포 일정: cols 14~19, '[...]' 또는 '공식블로그' 를 그룹 구분자로
    dist = []
    group = ""
    for r in sorted(byrow):
        row = byrow[r]
        head = str(row.get(14, "")).strip()
        if not head:
            continue
        if head.startswith("["):
            group = head.strip("[]").split("]")[0].strip()
            continue
        if head == "블로거":
            continue
        if head == "공식블로그":
            group = "공식블로그"
            continue
        pub = row.get(15)
        pub_s = _dstr(pub) if _is_date(pub) else ""
        note_extra = "" if (_is_date(pub) or pub is None) else str(pub).strip()
        dist.append({
            "group": group, "blogger": head, "publish_date": pub_s,
            "approval_no": str(row.get(16, "")).strip(),
            "landing_url": str(row.get(17, "")).strip(),
            "note": (note_extra or str(row.get(18, "")).strip()),
            "publish_url": str(row.get(19, "")).strip(),
        })
    return events, dist
